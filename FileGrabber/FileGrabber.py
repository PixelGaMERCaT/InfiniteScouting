from openpyxl import Workbook, load_workbook
import os
adbLocation = "adb" #"\"C:/Users/Cheesy Boi Tim/AppData/Local/Android/Sdk/platform-tools/adb\""
stream = os.popen(adbLocation + " devices -l ")
output = stream.readlines()
cwd = os.path.abspath(os.getcwd())
fullDataString = ""
for i in range(len(output) - 2):
    model = output[i + 1].split()[3].split(":")[1]
    id = output[i + 1].split()[0]
    print(id, model)
    cmd = adbLocation + " -s " + id + " pull "
    if (model == "Studio_5_0_S_II"):
        cmd += "/storage/sdcard0/Download/export.csv "
    elif (model == "N9130"):
        cmd += "/storage/emulated/legacy/Download/export.csv "
    else:
        cmd += "/storage/emulated/0/Download/export.csv "
    cmd += "\"" + cwd + "\\export" + str(i) + ".csv" + "\""
    print(cmd)
    a = os.popen(cmd)
    print(a.readlines())
    with open("export" + str(i) + ".csv", "r") as f:
        fullDataString += f.read()
data = []
rows = fullDataString.split("\n")

for i in rows:
    data.append(i.split(","))
del data[-1]

with open("export.csv", "w+") as f:
    f.write(fullDataString)

existingMatches = []
existingTeams = []

try:
    wb = load_workbook("data.xlsx")
    calculations = wb["Calculations"]
    dataSheet = wb["data"]
    for row in dataSheet.iter_rows(min_row = 2):
        existingMatches.append((int(row[1].value), int(row[2].value)))
    for row in calculations.iter_rows(min_row = 2):
        existingTeams.append(int(row[0].value))
except FileNotFoundError:
    wb = Workbook()
    calculations = wb.active
    calculations.title = "Calculations"
    dataSheet = wb.create_sheet("data")
    dataSheet.append(("Name", "Match Number", "Team Number", "Initation crossed", "Auto Inner", "Auto High", "Auto Low", "Teleop Inner", "Teleop High", "Teleop Low", "Defense Played", "Defense Effectiveness", "Defense Played Against", "Climb Effectiveness", "Comments"))
    calculations.append(("Team Number", "Matches", "% Leave initiation", "Avg Auto Inner", "Avg Auto High", "Avg Auto Low", "Avg Auto Score", "Avg Teleop Inner", "Avg Teleop High", "Avg Teleop Low", "Position Control", "Rotation Control", "% Rendevous", "% Times Climbed", "Avg Teleop Score", "% Times Played Defense", "Avg Defense Effectiveness", "Avg Score"))

#removing duplicates
duplicatesRemoved = []
for i in range(len(data)):
    data[i][1] = int(data[i][1])
    data[i][2] = int(data[i][2])
    if not (data[i][1], data[i][2]) in existingMatches:
        duplicatesRemoved.append(data[i])
        existingMatches.append((data[i][1], data[i][2]))
data = sorted(duplicatesRemoved, key=lambda x:x[1])

#adding to Calculations
for i in range(len(data)):
    if not data[i][1] in existingTeams:
        rowNum = calculations.max_row + 1
        rowCalcs = (data[i][2],                                                                                                                                                             #Team number
                    '=COUNTIF(Data.C2:C1000,A%d)' % rowNum,                                                                                                                #Number of matches
                    '=IFERROR(TRUNC(COUNTIFS($Data.$C$2:$C1000,$A%d,$Data.$D$2:$D1000,"=1") /$B%d * 100, 2),"")' % (rowNum, rowNum),                                                  #% Leave initiation
                    '=IFERRR(TRUNC(SUMIF($Data.$C$2:$C1000,$A%d,$Data.$F$2:$F1000)/$B%d,2),"")' % (rowNum, rowNum),                                                                  #Avg auto inner
                    '=IFERROR(TRUNC(SUMIF($Data.$C$2:$C1000,$A%d,$Data.$G$2:$G1000)/$B%d,2),"")' % (rowNum, rowNum),                                                                  #Avg auto high
                    '=IFERROR(TRUNC(SUMIF($Data.$C$2:$C1000,$A%d,$Data.$H$2:$H1000)/$B%d,2),"")' % (rowNum, rowNum),                                                                  #Avg auto low
                    '=(D%d*6+E%d*4+F%d*2+C%d/20)' % (rowNum, rowNum, rowNum, rowNum),                                                                                                                      #Avg auto score
                    '=IFERROR(TRUNC(SUMIF($Data.$C$2:$C1000,$A%d,$Data.$G$2:$G1000)/$B%d,2),"")' % (rowNum, rowNum),                                                                  #Avg teleop inner
                    '=IFERROR(TRUNC(SUMIF($Data.$C$2:$C1000,$A%d,$Data.$H$2:$H1000)/$B%d,2),"")' % (rowNum, rowNum),                                                                  #Avg teleop high
                    '=IFERROR(TRUNC(SUMIF($Data.$C$2:$C1000,$A%d,$Data.$I$2:$I1000)/$B%d,2),"")' % (rowNum, rowNum),                                                                  #Avg teleop low
                    '',                                                                                                                                                                     #Position Control
                    '',                                                                                                                                                                     #Rotation Control
                    '=IFERROR(TRUNC((1-COUNTIFS($Data.$C$2:$C1000,$A%d,$Data.$L$%d:$L1000,"=1") /$B2) * 100, 2),"")' % (rowNum, rowNum),                                              #% Rendevous
                    '=IFERROR(TRUNC(COUNTIFS($Data.$C$2:$C1000,$A%d,$Data.$L$%d:$L1000,"=3") /$B2 * 100, 2),"")' % (rowNum, rowNum),                                                  #% Times climbed
                    '=(H%d+I%d*2+J%d*3+M%d/20+N%d/20)' % (rowNum, rowNum, rowNum, rowNum, rowNum),                                                                                                                                                            #Avg teleop score
                    '=IFERROR(TRUNC(COUNTIFS($Data.$C$2:$C1000,$A%d,$Data.$M$%d:$M1000,"=1") /$B2 * 100, 2),"")' % (rowNum, rowNum),                                                  #% Times played defense
                    '=TRUNC(SUMIF($Data.$C$2:$C1000,$A%d,$Data.$N$%d:$N1000)/COUNTIFS($Data.$C$2:$C1000,$A%d,$Data.$M$%d:$M1000,"=1"), 2)' % (rowNum, rowNum, rowNum, rowNum),  #Avg defense score
                    '=(G%d+O%d)'                                                                                                                                                             #Avg overall score
                    )
        calculations.append(rowCalcs)


for row in data:
    dataSheet.append(row)

wb.save("data.xlsx")